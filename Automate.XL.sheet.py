

import openpyxl as xl, BarChart, Reference


def load_workbook(file_name):
    wb = xl.load_workbook(file_name)
    sheet = wb["sheet1"]

    for row in range (2,sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell= sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    wb.save(file_name)

    values = Reference(
        min_row = 2, max_row = sheet.max_row, min_col = 4, max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,)